#!/usr/bin/env python3
"""
TechFreedom — PocketBase Collection Setup & Data Import / Static JSON Export

Creates all collections with proper schemas and API rules,
then imports tools, alternatives, archetypes, and scoring guide data.

Usage (PocketBase import):
    pip3 install openpyxl requests
    python3 import-data.py --url https://api.techfreedom.eu --email admin@techfreedom.eu --password YOUR_PASSWORD --xlsx techfreedom-database.xlsx

Usage (static JSON export — no PocketBase needed):
    pip3 install openpyxl
    python3 import-data.py --xlsx techfreedom-database.xlsx --export-json assess/data/

Prerequisites (PocketBase import only):
    1. PocketBase running and accessible
    2. Admin account created via the PocketBase UI (https://api.techfreedom.eu/_/)
"""

import argparse
import json
import sys

try:
    import requests
except ImportError:
    requests = None  # Only needed for PocketBase import mode


# ============================================================
# Collection Schemas
# ============================================================

COLLECTIONS = [
    {
        "name": "tools",
        "type": "base",
        "schema": [
            {"name": "name", "type": "text", "required": True, "options": {"min": 1, "max": 200}},
            {"name": "slug", "type": "text", "required": True, "options": {"min": 1, "max": 100, "pattern": "^[a-z0-9-]+$"}},
            {"name": "category", "type": "text", "required": True},
            {"name": "provider", "type": "text", "required": True},
            {"name": "hqCountry", "type": "text"},
            {"name": "dataHosting", "type": "text"},
            {"name": "jurisdiction", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "continuity", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "surveillance", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "lockIn", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "costExposure", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "total", "type": "number"},
            {"name": "riskLevel", "type": "select", "options": {"values": ["Low", "Medium", "High", "Critical"]}},
            {"name": "keyRisks", "type": "editor"},
            {"name": "lastReviewed", "type": "text"},
        ],
        "listRule": "",       # Public read
        "viewRule": "",       # Public read
        "createRule": None,   # Admin only
        "updateRule": None,   # Admin only
        "deleteRule": None,   # Admin only
    },
    {
        "name": "archetypes",
        "type": "base",
        "schema": [
            {"name": "name", "type": "text", "required": True},
            {"name": "slug", "type": "text", "required": True, "options": {"min": 1, "max": 100, "pattern": "^[a-z0-9-]+$"}},
            {"name": "description", "type": "text"},
            {"name": "tools", "type": "relation", "required": True, "options": {"collectionId": "__tools__", "maxSelect": None}},
        ],
        "listRule": "",
        "viewRule": "",
        "createRule": None,
        "updateRule": None,
        "deleteRule": None,
    },
    {
        "name": "assessments",
        "type": "base",
        "schema": [
            {"name": "tools", "type": "relation", "options": {"collectionId": "__tools__", "maxSelect": None}},
            {"name": "scores", "type": "json"},
            {"name": "summary", "type": "editor"},
            {"name": "email", "type": "email"},
            {"name": "shareCode", "type": "text", "options": {"min": 0, "max": 50}},
        ],
        "listRule": None,     # No public list
        "viewRule": None,     # No public view (could allow by shareCode later)
        "createRule": "",     # Anyone can create
        "updateRule": None,
        "deleteRule": None,
    },
    {
        "name": "cohorts",
        "type": "base",
        "schema": [
            {"name": "name", "type": "text", "required": True},
            {"name": "startDate", "type": "date"},
            {"name": "lumaUrl", "type": "url"},
            {"name": "status", "type": "select", "options": {"values": ["Open", "Full", "Completed"]}},
            {"name": "capacity", "type": "number"},
        ],
        "listRule": "",
        "viewRule": "",
        "createRule": None,
        "updateRule": None,
        "deleteRule": None,
    },
    {
        "name": "scoring_guide",
        "type": "base",
        "schema": [
            {"name": "dimension", "type": "text", "required": True},
            {"name": "score", "type": "number", "required": True, "options": {"min": 1, "max": 5}},
            {"name": "label", "type": "text", "required": True},
            {"name": "description", "type": "editor"},
        ],
        "listRule": "",
        "viewRule": "",
        "createRule": None,
        "updateRule": None,
        "deleteRule": None,
    },
    {
        "name": "alternatives",
        "type": "base",
        "schema": [
            {"name": "name", "type": "text", "required": True},
            {"name": "slug", "type": "text", "required": True},
            {"name": "category", "type": "text", "required": True},
            {"name": "alternativeTo", "type": "text"},
            {"name": "provider", "type": "text"},
            {"name": "hqCountry", "type": "text"},
            {"name": "openSource", "type": "select", "options": {"values": ["Yes", "No", "Partially"]}},
            {"name": "selfHostable", "type": "select", "options": {"values": ["Yes", "No", "N/A"]}},
            {"name": "dataHosting", "type": "text"},
            {"name": "jurisdiction", "type": "number", "options": {"min": 1, "max": 5}},
            {"name": "continuity", "type": "number", "options": {"min": 1, "max": 5}},
            {"name": "surveillance", "type": "number", "options": {"min": 1, "max": 5}},
            {"name": "lockIn", "type": "number", "options": {"min": 1, "max": 5}},
            {"name": "costExposure", "type": "number", "options": {"min": 1, "max": 5}},
            {"name": "total", "type": "number"},
            {"name": "approxCost", "type": "text"},
            {"name": "migrationDifficulty", "type": "select", "options": {"values": ["Low", "Low-Medium", "Medium", "High"]}},
            {"name": "tradeoffs", "type": "editor"},
            {"name": "lastReviewed", "type": "text"},
        ],
        "listRule": "@request.auth.id != ''",  # Authenticated only
        "viewRule": "@request.auth.id != ''",  # Authenticated only
        "createRule": None,
        "updateRule": None,
        "deleteRule": None,
    },
]


# ============================================================
# Archetype Definitions
# ============================================================

ARCHETYPES = [
    {
        "name": "Microsoft Heavy",
        "slug": "microsoft-heavy",
        "description": "Typical organisation running on the Microsoft ecosystem",
        "tool_slugs": ["microsoft-365", "microsoft-teams", "linkedin", "dropbox", "eventbrite"],
    },
    {
        "name": "Google Heavy",
        "slug": "google-heavy",
        "description": "Organisation built around Google's tools and platforms",
        "tool_slugs": ["google-workspace", "gmail-free", "google-forms", "canva", "meta"],
    },
    {
        "name": "Typical Small Charity",
        "slug": "typical-small-charity",
        "description": "Common stack for small UK charities and community organisations",
        "tool_slugs": ["google-workspace", "canva", "mailchimp", "trello", "zoom", "whatsapp"],
    },
    {
        "name": "Startup",
        "slug": "startup",
        "description": "Fast-moving startup or social enterprise tech stack",
        "tool_slugs": ["slack", "aws", "hubspot", "asana", "calendly", "zoom"],
    },
    {
        "name": "AI Explorer",
        "slug": "ai-explorer",
        "description": "Organisation heavily integrating AI and cloud services",
        "tool_slugs": ["google-workspace", "slack", "aws", "zoom", "monday-com"],
    },
    {
        "name": "Legacy Stalwarts",
        "slug": "legacy-stalwarts",
        "description": "Established organisation with deep enterprise tool commitments",
        "tool_slugs": ["microsoft-365", "salesforce", "surveymonkey", "eventbrite", "wordpress-com"],
    },
]


# ============================================================
# Scoring Guide Data
# ============================================================

SCORING_GUIDE = [
    {"dimension": "jurisdiction", "score": 1, "label": "Minimal", "description": "UK/EU hosted. UK/EU company. GDPR compliant. Clear data residency."},
    {"dimension": "jurisdiction", "score": 2, "label": "Low", "description": "EU company with EU hosting. GDPR compliant. Minor jurisdictional questions."},
    {"dimension": "jurisdiction", "score": 3, "label": "Moderate", "description": "US company claiming GDPR compliance. EU hosting option available. Data location partly configurable."},
    {"dimension": "jurisdiction", "score": 4, "label": "High", "description": "US company. Data primarily in US. CLOUD Act applies. GDPR compliance questionable."},
    {"dimension": "jurisdiction", "score": 5, "label": "Critical", "description": "US company. Data in US. Subject to CLOUD Act & FISA. No data residency. Conflicts with GDPR."},
    {"dimension": "continuity", "score": 1, "label": "Minimal", "description": "Full data export in open formats. Multiple alternatives. Easy to switch. Self-hostable."},
    {"dimension": "continuity", "score": 2, "label": "Low", "description": "Good export options. Several alternatives. Switching manageable. Active community."},
    {"dimension": "continuity", "score": 3, "label": "Moderate", "description": "Export possible but painful. Some alternatives exist. Switching requires planning."},
    {"dimension": "continuity", "score": 4, "label": "High", "description": "Export difficult or incomplete. Few viable alternatives. Switching is a major project."},
    {"dimension": "continuity", "score": 5, "label": "Critical", "description": "No meaningful data export. No realistic alternative. Total vendor dependency."},
    {"dimension": "surveillance", "score": 1, "label": "Minimal", "description": "Open source. No tracking. Privacy-first design. No data harvesting."},
    {"dimension": "surveillance", "score": 2, "label": "Low", "description": "Minimal tracking. Opt-out available. Data not sold. Transparent privacy policy."},
    {"dimension": "surveillance", "score": 3, "label": "Moderate", "description": "Some tracking for product improvement. Opt-out partially available. Privacy policy complex."},
    {"dimension": "surveillance", "score": 4, "label": "High", "description": "Extensive tracking. Data used for ads or AI training. Opt-out limited. Metadata harvested."},
    {"dimension": "surveillance", "score": 5, "label": "Critical", "description": "Business model IS surveillance. Data sold or used for ads/AI. No opt-out. Pervasive tracking."},
    {"dimension": "lockIn", "score": 1, "label": "Minimal", "description": "Open standards. Full data portability. No proprietary formats. Easy to leave."},
    {"dimension": "lockIn", "score": 2, "label": "Low", "description": "Mostly open formats. Some proprietary features. Migration possible with effort."},
    {"dimension": "lockIn", "score": 3, "label": "Moderate", "description": "Mix of open and proprietary. Workflows create soft lock-in. Switching costs moderate."},
    {"dimension": "lockIn", "score": 4, "label": "High", "description": "Proprietary formats dominant. Deep workflow integration. Significant switching costs."},
    {"dimension": "lockIn", "score": 5, "label": "Critical", "description": "Totally proprietary. No data portability. Migration nearly impossible."},
    {"dimension": "costExposure", "score": 1, "label": "Minimal", "description": "Free/open source or stable pricing. No sharp increases. Multiple competitors."},
    {"dimension": "costExposure", "score": 2, "label": "Low", "description": "Affordable. Pricing stable historically. Free tier available. Competitive market."},
    {"dimension": "costExposure", "score": 3, "label": "Moderate", "description": "Mid-range pricing. Some price increases historically. Free tier limited."},
    {"dimension": "costExposure", "score": 4, "label": "High", "description": "Expensive per-seat. History of price increases. Free tier being eroded. Market dominance."},
    {"dimension": "costExposure", "score": 5, "label": "Critical", "description": "Monopoly pricing. Dramatic recent increases. Free tier eliminated. Captive audience."},
]


# ============================================================
# Helpers
# ============================================================

def slugify(name):
    import re
    s = name.lower().strip()
    # Handle special cases
    s = s.replace("microsoft 365", "microsoft-365")
    s = s.replace("monday.com", "monday-com")
    s = s.replace("wordpress.com", "wordpress-com")
    s = s.replace("wordpress.org", "wordpress-org")
    s = s.replace("cal.com", "cal-com")
    s = s.replace("x / twitter", "x-twitter")
    s = s.replace("gmail (free/personal)", "gmail-free")
    s = s.replace("whatsapp (organisational use)", "whatsapp")
    s = s.replace("meta (facebook / instagram)", "meta")
    # Generic cleanup
    s = re.sub(r'\s*\(.*?\)', '', s)  # Remove parenthetical
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = s.strip('-')
    return s


def safe_int(v):
    """Safely convert a value to int, returning 0 on failure."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def parse_bool_select(v):
    """Parse Yes/No/Partially/N-A style values from xlsx."""
    s = str(v or "").lower().strip()
    if s in ("yes", "true"):
        return "Yes"
    if s in ("no", "false"):
        return "No"
    if s in ("partially",):
        return "Partially"
    if s in ("n/a", "na"):
        return "N/A"
    return "No"


class PocketBaseClient:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip('/')
        self.token = None
        self.session = requests.Session()

    def authenticate(self, email, password):
        """Authenticate as admin."""
        resp = self.session.post(
            f"{self.base_url}/api/admins/auth-with-password",
            json={"identity": email, "password": password},
        )
        if resp.status_code != 200:
            print(f"Auth failed: {resp.status_code} {resp.text}")
            sys.exit(1)
        self.token = resp.json()["token"]
        self.session.headers["Authorization"] = self.token
        print(f"  Authenticated as {email}")

    def list_collections(self):
        """List existing collections."""
        resp = self.session.get(f"{self.base_url}/api/collections")
        resp.raise_for_status()
        return {c["name"]: c for c in resp.json().get("items", resp.json() if isinstance(resp.json(), list) else [])}

    def get_collection(self, name):
        """Get a collection by name."""
        resp = self.session.get(f"{self.base_url}/api/collections/{name}")
        if resp.status_code == 404:
            return None
        resp.raise_for_status()
        return resp.json()

    def create_collection(self, definition):
        """Create a collection."""
        resp = self.session.post(
            f"{self.base_url}/api/collections",
            json=definition,
        )
        if resp.status_code not in (200, 201):
            print(f"  Failed to create collection '{definition['name']}': {resp.status_code}")
            print(f"  {resp.text}")
            return None
        return resp.json()

    def create_record(self, collection, data):
        """Create a record in a collection."""
        resp = self.session.post(
            f"{self.base_url}/api/collections/{collection}/records",
            json=data,
        )
        if resp.status_code not in (200, 201):
            print(f"  Failed to create record in '{collection}': {resp.status_code}")
            print(f"  {resp.text[:200]}")
            return None
        return resp.json()

    def list_records(self, collection, per_page=200, filter_str=None):
        """List records in a collection."""
        params = {"perPage": per_page}
        if filter_str:
            params["filter"] = filter_str
        resp = self.session.get(
            f"{self.base_url}/api/collections/{collection}/records",
            params=params,
        )
        resp.raise_for_status()
        return resp.json().get("items", [])


# ============================================================
# Main
# ============================================================

def read_tools_from_xlsx(xlsx_path):
    """Read tools from xlsx and return list of dicts in internal format."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_tools = wb[wb.sheetnames[0]]
    tools = []

    for i, row in enumerate(ws_tools.iter_rows(min_row=2)):
        vals = [c.value for c in row]
        if vals[0] is None:
            break

        name = str(vals[0])
        slug = slugify(name)

        tools.append({
            "id": i + 1,
            "name": name,
            "slug": slug,
            "category": str(vals[1] or ""),
            "provider": str(vals[2] or ""),
            "hqCountry": str(vals[3] or ""),
            "dataHosting": str(vals[4] or ""),
            "jurisdiction": safe_int(vals[5]),
            "continuity": safe_int(vals[6]),
            "surveillance": safe_int(vals[7]),
            "lockIn": safe_int(vals[8]),
            "costExposure": safe_int(vals[9]),
            "total": safe_int(vals[10]),
            "riskLevel": str(vals[11] or ""),
            "keyRisks": str(vals[12] or ""),
        })

    return tools


def build_archetypes(tools):
    """Build archetypes with toolSlugs referencing tools by slug."""
    tool_slugs_set = {t["slug"] for t in tools}
    archetypes = []

    for arch in ARCHETYPES:
        tool_slugs = [s for s in arch["tool_slugs"] if s in tool_slugs_set]
        archetypes.append({
            "name": arch["name"],
            "slug": arch["slug"],
            "description": arch["description"],
            "toolSlugs": tool_slugs,
        })

    return archetypes


def read_alternatives_from_xlsx(xlsx_path):
    """Read alternatives from xlsx sheet[1] and return list of dicts in internal format."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_alts = wb[wb.sheetnames[1]]
    alternatives = []

    for i, row in enumerate(ws_alts.iter_rows(min_row=2)):
        vals = [c.value for c in row]
        if vals[0] is None:
            break

        name = str(vals[0])
        slug = slugify(name)

        # alternativeTo: split on commas and slugify each part
        alt_to_raw = str(vals[2] or "")
        alternative_to = [slugify(part.strip()) for part in alt_to_raw.split(",") if part.strip()]

        alternatives.append({
            "id": i + 1,
            "name": name,
            "slug": slug,
            "category": str(vals[1] or ""),
            "alternativeTo": alternative_to,
            "provider": str(vals[3] or ""),
            "hqCountry": str(vals[4] or ""),
            "openSource": parse_bool_select(vals[5]),
            "selfHostable": parse_bool_select(vals[6]),
            "dataHosting": str(vals[7] or ""),
            "jurisdiction": safe_int(vals[8]),
            "continuity": safe_int(vals[9]),
            "surveillance": safe_int(vals[10]),
            "lockIn": safe_int(vals[11]),
            "costExposure": safe_int(vals[12]),
            "total": safe_int(vals[13]),
            "approxCost": str(vals[14] or ""),
            "migrationDifficulty": str(vals[15] or ""),
            "tradeoffs": str(vals[16] or ""),
            "lastReviewed": str(vals[17] or ""),
        })

    return alternatives


def export_json(xlsx_path, output_dir):
    """Export tools.json and archetypes.json to output_dir."""
    import os

    os.makedirs(output_dir, exist_ok=True)

    print("\nReading xlsx...")
    tools = read_tools_from_xlsx(xlsx_path)
    print(f"  Found {len(tools)} tools")

    archetypes = build_archetypes(tools)
    print(f"  Built {len(archetypes)} archetypes")

    alternatives = read_alternatives_from_xlsx(xlsx_path)
    print(f"  Found {len(alternatives)} alternatives")

    tools_path = os.path.join(output_dir, "tools.json")
    with open(tools_path, "w", encoding="utf-8") as f:
        json.dump(tools, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {tools_path}")

    archetypes_path = os.path.join(output_dir, "archetypes.json")
    with open(archetypes_path, "w", encoding="utf-8") as f:
        json.dump(archetypes, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {archetypes_path}")

    alternatives_path = os.path.join(output_dir, "alternatives.json")
    with open(alternatives_path, "w", encoding="utf-8") as f:
        json.dump(alternatives, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {alternatives_path}")

    print("\nDone! Static JSON files ready for deployment.")


def main():
    parser = argparse.ArgumentParser(description="TechFreedom PocketBase setup & data import")
    parser.add_argument("--url", help="PocketBase URL (e.g. https://api.techfreedom.eu)")
    parser.add_argument("--email", help="Admin email")
    parser.add_argument("--password", help="Admin password")
    parser.add_argument("--xlsx", required=True, help="Path to techfreedom-database.xlsx")
    parser.add_argument("--export-json", metavar="DIR", help="Export tools.json and archetypes.json to DIR (no PocketBase needed)")
    parser.add_argument("--skip-collections", action="store_true", help="Skip collection creation")
    parser.add_argument("--skip-import", action="store_true", help="Skip data import")
    args = parser.parse_args()

    # ---- Static JSON export mode ----
    if args.export_json:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
            sys.exit(1)
        export_json(args.xlsx, args.export_json)
        return

    # ---- PocketBase import mode — requires credentials ----
    if not args.url or not args.email or not args.password:
        parser.error("--url, --email, and --password are required for PocketBase import (or use --export-json)")

    pb = PocketBaseClient(args.url)

    # ---- Authenticate ----
    print("\n[1/5] Authenticating...")
    pb.authenticate(args.email, args.password)

    # ---- Create Collections ----
    if not args.skip_collections:
        print("\n[2/5] Creating collections...")
        collection_ids = {}

        for coll_def in COLLECTIONS:
            name = coll_def["name"]
            existing = pb.get_collection(name)
            if existing:
                print(f"  '{name}' already exists, skipping")
                collection_ids[name] = existing["id"]
                continue

            # Build the schema, replacing __tools__ placeholder with actual ID
            schema = []
            for field in coll_def["schema"]:
                field_copy = dict(field)
                if field_copy.get("options", {}).get("collectionId") == "__tools__":
                    if "tools" not in collection_ids:
                        print(f"  ERROR: 'tools' collection must be created before '{name}'")
                        sys.exit(1)
                    field_copy["options"] = dict(field_copy["options"])
                    field_copy["options"]["collectionId"] = collection_ids["tools"]
                schema.append(field_copy)

            payload = {
                "name": name,
                "type": coll_def["type"],
                "schema": schema,
                "listRule": coll_def.get("listRule"),
                "viewRule": coll_def.get("viewRule"),
                "createRule": coll_def.get("createRule"),
                "updateRule": coll_def.get("updateRule"),
                "deleteRule": coll_def.get("deleteRule"),
            }

            result = pb.create_collection(payload)
            if result:
                collection_ids[name] = result["id"]
                print(f"  Created '{name}' (id: {result['id']})")
            else:
                sys.exit(1)
    else:
        print("\n[2/5] Skipping collection creation")

    # ---- Import Data ----
    if not args.skip_import:
        print("\n[3/5] Reading xlsx data...")
        try:
            import openpyxl
        except ImportError:
            print("  ERROR: openpyxl not installed. Run: pip3 install openpyxl")
            sys.exit(1)

        wb = openpyxl.load_workbook(args.xlsx, data_only=True)
        sheets = wb.sheetnames

        # -- Tools --
        print("\n[4/5] Importing tools...")
        ws_tools = wb[sheets[0]]
        tool_slug_to_id = {}

        for row in ws_tools.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if vals[0] is None:
                break

            name = str(vals[0])
            slug = slugify(name)

            data = {
                "name": name,
                "slug": slug,
                "category": str(vals[1] or ""),
                "provider": str(vals[2] or ""),
                "hqCountry": str(vals[3] or ""),
                "dataHosting": str(vals[4] or ""),
                "jurisdiction": safe_int(vals[5]),
                "continuity": safe_int(vals[6]),
                "surveillance": safe_int(vals[7]),
                "lockIn": safe_int(vals[8]),
                "costExposure": safe_int(vals[9]),
                "total": safe_int(vals[10]),
                "riskLevel": str(vals[11] or ""),
                "keyRisks": str(vals[12] or ""),
                "lastReviewed": str(vals[13] or ""),
            }

            record = pb.create_record("tools", data)
            if record:
                tool_slug_to_id[slug] = record["id"]
                print(f"  + {name} ({slug})")

        print(f"  Imported {len(tool_slug_to_id)} tools")

        # -- Alternatives --
        print("\n  Importing alternatives...")
        ws_alts = wb[sheets[1]]
        alt_count = 0

        for row in ws_alts.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if vals[0] is None:
                break

            name = str(vals[0])
            slug = slugify(name)

            # Slugify alternativeTo as comma-separated slugs for consistent references
            alt_to_raw = str(vals[2] or "")
            alt_to_slugs = ",".join(slugify(part.strip()) for part in alt_to_raw.split(",") if part.strip())

            data = {
                "name": name,
                "slug": slug,
                "category": str(vals[1] or ""),
                "alternativeTo": alt_to_slugs,
                "provider": str(vals[3] or ""),
                "hqCountry": str(vals[4] or ""),
                "openSource": parse_bool_select(vals[5]),
                "selfHostable": parse_bool_select(vals[6]),
                "dataHosting": str(vals[7] or ""),
                "jurisdiction": safe_int(vals[8]),
                "continuity": safe_int(vals[9]),
                "surveillance": safe_int(vals[10]),
                "lockIn": safe_int(vals[11]),
                "costExposure": safe_int(vals[12]),
                "total": safe_int(vals[13]),
                "approxCost": str(vals[14] or ""),
                "migrationDifficulty": str(vals[15] or ""),
                "tradeoffs": str(vals[16] or ""),
                "lastReviewed": str(vals[17] or ""),
            }

            record = pb.create_record("alternatives", data)
            if record:
                alt_count += 1
                print(f"  + {name}")

        print(f"  Imported {alt_count} alternatives")

        # -- Archetypes --
        print("\n  Importing archetypes...")
        for arch in ARCHETYPES:
            tool_ids = []
            for ts in arch["tool_slugs"]:
                if ts in tool_slug_to_id:
                    tool_ids.append(tool_slug_to_id[ts])
                else:
                    print(f"  WARNING: Tool slug '{ts}' not found for archetype '{arch['name']}'")

            data = {
                "name": arch["name"],
                "slug": arch["slug"],
                "description": arch["description"],
                "tools": tool_ids,
            }
            record = pb.create_record("archetypes", data)
            if record:
                print(f"  + {arch['name']} ({len(tool_ids)} tools)")

        # -- Scoring Guide --
        print("\n  Importing scoring guide...")
        for entry in SCORING_GUIDE:
            pb.create_record("scoring_guide", entry)
        print(f"  Imported {len(SCORING_GUIDE)} scoring guide entries")

    else:
        print("\n[3-4/5] Skipping data import")

    # ---- Summary ----
    print("\n[5/5] Done!")
    print(f"\n  Admin UI: {args.url}/_/")
    print(f"  API: {args.url}/api/collections/tools/records")
    print(f"  API: {args.url}/api/collections/archetypes/records?expand=tools")
    print()


if __name__ == "__main__":
    main()
